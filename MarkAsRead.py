# -*- coding: utf-8 -*-
"""
Created on Sun May 12 05:10:04 2019

@author: kpmg
"""

# -*- coding: utf-8 -*-
"""
Created on Sat May 11 13:06:02 2019

@author: kpmg
"""

import openpyxl
import os
import pandas as pd
import time
import datetime
import re
import sys
import win32com.client
from datetime import *
import os

#import scipy.interpolate.interpnd

    
def markAsRead():
    
    outlook=win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox=outlook.GetDefaultFolder(6)
    messages=inbox.Items
    
    
    wb = openpyxl.load_workbook("DashBoard.xlsm")
#    sheet_Time = wb["TimeSorted"]
#    sheet_Address = wb["addressedToYou"]
#    sheet_Designation = wb["DesignationSorted"]
#    sheet_Type = wb["TypeOfEmail"]
#    sheet_Negative = wb["isNegativeContext"]
#    sheet_Priority = wb["overallPriority"]
    sheet_getCloud = wb["getWordCloud"]
    
    for r in range(2, sheet_getCloud.max_row+1):
        mark = sheet_getCloud.cell(row=r, column=12).value
        if(mark=="No"):
            continue
        time = sheet_getCloud.cell(row=r, column=8).value
        subject = sheet_getCloud.cell(row=r, column=5).value
        
        if mark == "Yes":
            
            for message in messages:
                c = 1
                if(c>100):
                    break
                if message.Unread != True:
                    continue
                c+=1
                if datetime.strftime(message.ReceivedTime ,'%Y, %b %d, %H:%M') == time:
                    if message.Subject == subject:
                        message.Unread = False

    
    
if __name__ == '__main__':
    
    os.chdir(r"C:\Users\kpmg\Desktop\Hackathon")
    markAsRead()
    