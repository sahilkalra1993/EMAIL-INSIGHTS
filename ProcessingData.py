# -*- coding: utf-8 -*-
"""
Created on Sat May 11 13:06:02 2019

@author: kpmg
"""

import openpyxl
import os
import pandas as pd
import time
import datetime
from wordcloud import WordCloud, STOPWORDS 
import matplotlib.pyplot as plt 
import re
from textblob import TextBlob 
import sys
#import scipy.interpolate.interpnd

    
def clean_data(text): 
        ''' 
        Utility function to clean tweet text by removing links, special characters 
        using simple regex statements. 
        '''
        return ' '.join(re.sub("(@[A-Za-z0-9]+)|([^0-9A-Za-z \t])|(\w+:\/\/\S+)", " ", text).split())
    

def get_sentiment(text):    
    ''' 
        Utility function to classify sentiment of passed tweet 
        using textblob's sentiment method 
    '''
    # create TextBlob object of passed tweet text 
    analysis = TextBlob(clean_data(text)) 
    # set sentiment 
    if analysis.sentiment.polarity > 0: 
        return 'positive', analysis.sentiment.polarity, analysis.sentiment.subjectivity 
    elif analysis.sentiment.polarity == 0: 
        return 'neutral', analysis.sentiment.polarity,  analysis.sentiment.subjectivity
    else: 
        return 'negative', analysis.sentiment.polarity,  analysis.sentiment.subjectivity
        

def show_wordcloud(data, title = None):
    stopwords = set(STOPWORDS)
    wordcloud = WordCloud(
        background_color='white',
        stopwords=stopwords,
        max_words=200,
        max_font_size=50, 
        scale=3,
        random_state=1 # chosen at random by flipping a coin; it was heads
    ).generate(str(data))

    fig = plt.figure(1, figsize=(12, 12))
    plt.axis('off')
    if title: 
        fig.suptitle(title, fontsize=20)
        fig.subplots_adjust(top=2.3)
    
    
#    plt.imshow(wordcloud)
#    plt.show()
    return wordcloud
    

def getOrgLevel():
    wb = openpyxl.load_workbook("OrgChart.xlsx")
    sheet = wb["OrgTree"]
    dict_position = {}
    for c in range(1,  sheet.max_column+1):
        for r in range(2, sheet.max_row+1):
            if sheet.cell(row=r, column = c).value:
                dict_position[sheet.cell(row=r, column = c).value] = sheet.cell(row=1, column = c).value
    
    return dict_position


def getTimePriority(hour_threshold = 18, minute_threshold = 0):
    
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    wb.create_sheet("TimeSorted")
    sheet_SortedTime = wb["TimeSorted"]
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_SortedTime.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value 
##    wb.create_sheet("SortedTime")
#    sheet_SortedTime = wb["SortedTime"]
    sheet_SortedTime.cell(row=1, column=10).value = "UNIX TIME"
    sheet_SortedTime.cell(row=1, column=11).value = "SHOW FLAG"
    
    for r in range(2, sheet_SortedTime.max_row+1):
        sheet_SortedTime.cell(row=r, column=10).value = int(time.time()) - int(time.mktime(datetime.datetime.strptime(sheet_SortedTime.cell(row=r, column=8).value, "%Y, %b %d, %H:%M").timetuple()))
        
        if int(sheet_SortedTime.cell(row=r, column=10).value) > (hour_threshold*60*60 + minute_threshold*60):
            sheet_SortedTime.cell(row=r, column=11).value = "Yes"
        else:
            sheet_SortedTime.cell(row=r, column=11).value = "No"
    wb.save("Email_dump.xlsx")
    

def getDesignationPriority():
    own_id = ''
    with open('Own_Email_ID.txt', 'r') as f:
        own_id = f.read()
    own_id = getOrgLevel()[own_id]
#    print ("OWN ID: ", own_id)
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    
    wb.create_sheet("DesignationSorted")
    sheet_DesignationSorted = wb["DesignationSorted"]
    
    sheet_DesignationSorted.cell(row=1, column=10).value = "DESIGNATION LEVEL"
#    sheet_DesignationSorted.cell(row=1, column=12).value = "SENDER DESIGNATION"
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_DesignationSorted.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value 
    
    for r in range(2, sheet_DesignationSorted.max_row+1):
        if sheet_DesignationSorted.cell(row=r, column=4).value in getOrgLevel():
#            print ("Receiving ID: ", int(getOrgLevel()[sheet_DesignationSorted.cell(row=r, column=4).value]))
            if int(getOrgLevel()[sheet_DesignationSorted.cell(row=r, column=4).value]) - int(own_id) < 0:
                sheet_DesignationSorted.cell(row=r, column=10).value = "Senior"
            elif int(getOrgLevel()[sheet_DesignationSorted.cell(row=r, column=4).value]) - int(own_id) == 0:
                sheet_DesignationSorted.cell(row=r, column=10).value = "Colleague"
            else:
                sheet_DesignationSorted.cell(row=r, column=10).value = "Junior"
        else:
            sheet_DesignationSorted.cell(row=r, column=10).value = "Outside Domain"
    wb.save("Email_dump.xlsx")
    

def addressedToYou():
    own_id = ''
    with open('Own_Email_ID.txt', 'r') as f:
        own_id = f.read()
    
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    
    wb.create_sheet("addressedToYou")
    sheet_addressedToYou = wb["addressedToYou"]
    
    sheet_addressedToYou.cell(row=1, column=10).value = "ADDRESSED TO YOU"
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_addressedToYou.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value 
            
    for r in range(2, sheet_addressedToYou.max_row+1):
        mail_to = sheet_addressedToYou.cell(row=r, column=1).value
        mail_to = mail_to.split(";")
        sheet_addressedToYou .cell(row=r, column=10).value = "No"
        if len(mail_to) == 1:
            mail_to = mail_to[0]
            if mail_to == own_id:
                sheet_addressedToYou .cell(row=r, column=10).value = "Yes"
                
    wb.save("Email_dump.xlsx")


def typeOfEmail():
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    
    wb.create_sheet("TypeOfEmail")
    sheet_TypeOfEmail = wb["TypeOfEmail"]
    
    sheet_TypeOfEmail.cell(row=1, column=10).value = "TYPE OF EMAIL"
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_TypeOfEmail.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value 
            
    for r in range(2, sheet_TypeOfEmail.max_row+1): 
        text = sheet_TypeOfEmail.cell(row=r, column=5).value
        text = text.split(":")
        sheet_TypeOfEmail.cell(row=r, column=10).value = "New Email"
        if len(text)>1:
            if(text[0] == 'Fw' or text[0] == 'Re'):
                if text[0] == 'Fw':
                    sheet_TypeOfEmail.cell(row=r, column=10).value = "Forward"
                elif text[0] == 'Re':
                    sheet_TypeOfEmail.cell(row=r, column=10).value = "Reply"
        
    wb.save("Email_dump.xlsx")
        


def getWordCloud():
    
    try:  
        os.mkdir(r"C:\Users\kpmg\Desktop\Hackathon\WordCloud")
    except OSError:  
        pass
    else:  
        pass
    
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    
    wb.create_sheet("getWordCloud")
    sheet_getWordCloud = wb["getWordCloud"]
    
    sheet_getWordCloud.cell(row=1, column=10).value = "LOCATION"
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_getWordCloud.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value 
    
    for r in range(2, sheet_getWordCloud.max_row+1): 
        print (r)
        text = sheet_getWordCloud.cell(row=r, column=6).value
        text = clean_data(text)
        subject = sheet_getWordCloud.cell(row=r, column=5).value[:40]
        wordCloud = show_wordcloud(text, subject)
        if wordCloud:
            wordCloud.to_file(r'C:\Users\kpmg\Desktop\Hackathon\WordCloud\{}.jpg'.format(r))
            sheet_getWordCloud.cell(row=r, column=10).value = r'C:\Users\kpmg\Desktop\Hackathon\WordCloud\{}.jpg'.format(r)
    wb.save("Email_dump.xlsx")
    
    
def isNegativeContext():
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    
    wb.create_sheet("isNegativeContext")
    sheet_isNegativeContext = wb["isNegativeContext"]
    
    sheet_isNegativeContext.cell(row=1, column=10).value = "IS NEGATIVE"
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_isNegativeContext.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value
            
    for r in range(2, sheet_isNegativeContext.max_row+1): 
        text = sheet_isNegativeContext.cell(row=r, column=5).value +" "+sheet_isNegativeContext.cell(row=r, column=6).value
#        print (text[:500])
        text = clean_data(text)
        
        sentiment, polarity, subjectivity = get_sentiment(text)
        print (r, sentiment, polarity, subjectivity)
        if(sentiment == "negative"):
            sheet_isNegativeContext.cell(row=r, column=10).value = "Yes"
        else:
            sheet_isNegativeContext.cell(row=r, column=10).value = "No"
        
        
    wb.save("Email_dump.xlsx")
    
    
def overallPriority():
    wb = openpyxl.load_workbook("Email_dump.xlsx")
    sheet = wb["InputData"] 
    
    wb.create_sheet("overallPriority")
    sheet_overallPriority = wb["overallPriority"]
    
    sheet_overallPriority.cell(row=1, column=10).value = "OVERALL PRIORITY"
    
    
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            sheet_overallPriority.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value
            
    sheet_Time = wb["TimeSorted"]
    sheet_Address = wb["addressedToYou"]
    sheet_Designation = wb["DesignationSorted"]
    sheet_Type = wb["TypeOfEmail"]
    sheet_Negative = wb["isNegativeContext"]
    
    for r in range(2, sheet_overallPriority.max_row+1):
        priority_value = 0
        priority_value+=int(sheet_overallPriority.cell(row=r, column=7).value)
        if int(sheet_Time.cell(row=r, column=10).value)/3600 > 12:
            priority_value+=1
        if int(sheet_Time.cell(row=r, column=10).value)/3600 > 24:
            priority_value+=1
        if sheet_Designation.cell(row=r, column=10).value == 'Senior':
            priority_value+=2
            if sheet_Address.cell(row=r, column=10).value == 'Yes':
                priority_value+=2
        if sheet_Address.cell(row=r, column=10).value == 'Yes':
            priority_value+=1
        if sheet_Type.cell(row=r, column=10).value == 'New Email':
            priority_value+=1
        if sheet_Negative.cell(row=r, column=10).value == 'Yes':
            priority_value+=3
            
        sheet_overallPriority.cell(row=r, column=10).value = priority_value
        
        
    wb.save("Email_dump.xlsx")
    
    
if __name__ == '__main__':
    
    print (sys.argv[1])
    os.chdir(r"{}".format(sys.argv[1]))
    stopwords = set(STOPWORDS)

        
    getTimePriority()
    getDesignationPriority()
    addressedToYou()
    typeOfEmail()
    getWordCloud()
    isNegativeContext()
    overallPriority()
    
    with open("Solution1_complete.txt", "w") as f:
        f.write("Successful")
        