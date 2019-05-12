# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

import win32com.client
from datetime import *
import os
import openpyxl


os.chdir(r"C:\Users\kpmg\Desktop\Hackathon")

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "InputData"

outlook=win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")

inbox=outlook.GetDefaultFolder(6)

messages=inbox.Items
count = 1

r = 2
#with open('maildata.txt', 'w', encoding="utf-8") as f:

sheet.cell(row=1, column=1).value = "TO"
sheet.cell(row=1, column=2).value = "CC"
sheet.cell(row=1, column=3).value = "BCC"
sheet.cell(row=1, column=4).value = "FROM"
sheet.cell(row=1, column=5).value = "SUBJECT"
sheet.cell(row=1, column=6).value = "BODY"
sheet.cell(row=1, column=7).value = "PRIORITY"
sheet.cell(row=1, column=8).value = "MAIL_RECEIVED_TIME"
sheet.cell(row=1, column=9).value = "ATTACHMENT"


for message in messages:
    c = 1
    if(count>100):
        break
    if message.Unread != True:
        continue
#    print(message.Subject)
##     print(message.Body)
#    print(message.To)
#    print(message.Recipients)
#    print(message.Sender)
#    print(message.Sender.Address)
#    print(message.Importance)
#    print(datetime.strftime(message.ReceivedTime ,'%Y %a %b %d %H:%M:%S '))
    
    
    sheet.cell(row=r, column=c).value = message.To
    c+=1
    sheet.cell(row=r, column=c).value = message.CC
    c+=1
    sheet.cell(row=r, column=c).value = message.BCC
    c+=1
    sheet.cell(row=r, column=c).value = message.Sender.Address
    c+=1
    sheet.cell(row=r, column=c).value = message.Subject
    c+=1
    sheet.cell(row=r, column=c).value = message.Body
    c+=1
    sheet.cell(row=r, column=c).value = message.Importance
    c+=1    
    sheet.cell(row=r, column=c).value = datetime.strftime(message.ReceivedTime ,'%Y, %b %d, %H:%M')
    c+=1
#    sheet.cell(row=r, column=c).value = "{}".format([i for i in message.Attachments]) if message.Attachments else False
    sheet.cell(row=r, column=c).value = message.Attachments.Count
#    sheet.cell(row=r, column=c).value = message.Attachments.Class
    c+=1
#    sheet.cell(row=r, column=c).value = message.Size/(1024)
#    c+=1
#    
    
    r+=1
    
wb.save("Email_dump.xlsx")
#    sheet.cell(row=r, column=c).value = message.ReceivedTime
#    c+=1    
    
##        f.write(message.Body)
#    f.write('\n\n========\n\n')
#    
##     message.Unread = True
#    count+=1
##     attachments = message.attachments
#
##     for attachment in attachments:
##         pass
